import openpyxl
import time
from bs4 import BeautifulSoup
import requests
#Make Excel file
wb = openpyxl.load_workbook('1000companydataset.xlsx',data_only=True)
ws = wb.active

# HTTP GET Request
for i in range(100):
    req = requests.get('http://m.mk.co.kr/yearbook/index.php?page='+str(i)+'&TM=Y2&MM=T0')
    #매경연감 페이지는 page= 뒤에 붙은 숫자가 바뀌면서 페이지가 이동하므로 숫자를 변수로 활용하여 Iterate

    #참조 : http://pythonstudy.xyz/python/article/403-%ED%8C%8C%EC%9D%B4%EC%8D%AC-Web-Scraping
    req.encoding=None

    # HTML 소스 가져오기
    html = req.text

    #text화 된 html 변수를 BS에 담아서 parsing 할 수 있게 함
    soup = BeautifulSoup(html, 'html.parser')

    #1. soup에서 tbody를 찾고, 그안에서 a tag 탐색

    tds = soup.find("tbody")
    companies=tds.find_all('a')
    # 2. soup에서 tbody중 td tag를 찾음. td 태그 중 class속성이 center인것을 탐색
    ranking=tds.find_all('td',{'class':'center'})

    #list value initialize
    company=[]
    ranks=[]

    for a in companies:
        #a 태그에서 web browser에 표시되는 텍스트만 추출하고, (주)를 없애버림
        b=a.text
        b=b.replace('(주)','')
        # company 리스트에 전처리된 b를 하나씩 삽입
        company.append(b)

    for rank in ranking:
        #ranks list에 rank변수의 text값만 삽입
        ranks.append(rank.text)

    #excel 파일위치마다 company 정보 삽입
    for j in range(10):
        #print(company[j])
        ws.cell(row=i*10+j+2, column=1).value=company[j]
        ws.cell(row=i*10+j+2, column=2).value=int(ranks[j])

    #저장
    wb.save('1000companydataset.xlsx')
    # 크롤링 탐지 회피를 위한 sleep
    time.sleep(3)
    print('page '+str(i)+' done.')