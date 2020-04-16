from bs4 import BeautifulSoup
import openpyxl
import requests
import time
wb = openpyxl.load_workbook('seoul_strong_small_companies.xlsx',data_only=True)
ws = wb.active
url='http://www.seouljobnow.co.kr/bbs/board.php?bo_table=company&page='
for i in range(1,26):
    req = requests.get(url + str(i))

    html = req.text

    soup = BeautifulSoup(html, 'html.parser')

    trs = soup.findAll('tr',{'align':'center'})
    for idx, tr in enumerate(trs):
        #print(tr)
        tds = tr.findAll('td')
        addr = tds[5].text
        print(tds[1].text, tds[3].text,end=' ')
        if '서울특별시 ' in addr:
            addr = addr.replace('서울특별시 ','')
        if '서울시 ' in addr:
            addr = addr.replace('서울시 ', '')
        if '서울 ' in addr:
            addr = addr.replace('서울 ', '')
        print(addr)
        spaceIdx = addr.find(' ')

        #회사명
        ws.cell(row= 500 - int(tds[0].text) + 2, column=1).value = tds[1].text
        #업종
        ws.cell(row= 500 - int(tds[0].text) + 2, column=2).value = tds[3].text
        #행정구(ex 강남구, 송파구)
        ws.cell(row=500 - int(tds[0].text) + 2, column=3).value = addr[:spaceIdx]
        #세부주소
        ws.cell(row=500 - int(tds[0].text) + 2, column=4).value = addr[spaceIdx+1:]

        if(int(tds[0].text) == 1):
            wb.save('seoul_strong_small_companies.xlsx')
            break
    wb.save('seoul_strong_small_companies.xlsx')
    print(i,' page saved ')

