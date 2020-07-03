import openpyxl
import requests
import time
from bs4 import BeautifulSoup
#잡플래닛 사이트를 긁어오는 것이라 엑셀파일은 원하는 대로 변경
wb = openpyxl.load_workbook('seoul_strong_small_companies.xlsx',data_only=True)
ws = wb.active
url='https://www.jobplanet.co.kr/search?utf8=&query='
# 엑셀파일 위치 조작
for idx,i in enumerate(ws['A2':'A501']):
    print(idx+1 ,i[0].value,end=' ')
    req = requests.get(url+i[0].value)

    html = req.text

    soup = BeautifulSoup(html, 'html.parser')
    #print(soup)
    try:
        result_card = soup.find('span',{'class':'rate_ty02'})
        print(result_card.text)
        ws.cell(row=i[0].row, column = 5).value = result_card.text
        if idx%10 == 0:
            wb.save('./seoul_strong_small_companies.xlsx')
            print('saved.')

    except Exception as e:
        print('avg not found')
    time.sleep(2)

wb.save('seoul_strong_small_companies.xlsx')





