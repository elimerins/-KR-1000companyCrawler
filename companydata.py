import xlrd
from selenium import webdriver
import openpyxl
from selenium.webdriver.common.keys import Keys
import time

chrome_path='/Users/joyeongmin/Documents/2019/companies/chromedriver'
driver = webdriver.Chrome(chrome_path)

driver.implicitly_wait(30)

wb = openpyxl.load_workbook('1000companydataset.xlsx',data_only=True)
ws = wb.active

urlname='https://kreditjob.com/'

for r in ws['A2':'H1011']:
    driver.get(urlname)
    company=driver.find_element_by_xpath('//*[@id="root"]/div/div[5]/div[1]/div/div[2]/div/div[2]/div/input')
    companyname = r[0].value
    print(r[0].row-1)
    print(companyname)
    company.send_keys(companyname)
    company.send_keys(Keys.RETURN)
    #url=company.find_element_by_xpath('//*[@id="root"]/div/div[2]/ul/a[4]').text
    time.sleep(4)
    if urlname!=driver.current_url:
        #평균임금
        avg_sal=driver.find_element_by_xpath('//*[@id="root"]/div/div[5]/div[1]/div[1]/div[2]/section[2]/div[3]/div[2]/div[2]/span[2]').text
        print(avg_sal)
        avg_sal=avg_sal.replace(',','')
        try:
            ws.cell(row=r[0].row, column=3).value =int(avg_sal)
        except Exception as inst:
            print(type(inst))
            ws.cell(row=r[0].row, column=3).value =avg_sal
            continue

        #대졸 평균임금
        try:
            degree = driver.find_element_by_xpath('//*[@id="root"]/div/div[5]/div[1]/div[1]/div[2]/section[4]/div[2]/div[3]/div[1]/div[3]/span').text
            salary = driver.find_element_by_xpath('//*[@id="root"]/div/div[5]/div[1]/div[1]/div[2]/section[4]/div[2]/div[3]/div[1]/div[3]/div[2]').text
            salary = salary.replace(',', '')
            salary = salary[3:len(salary) - 2]

            print(degree,salary)
            ws.cell(row=r[0].row, column=4).value = int(salary)#1억이 넘는경우 3:8로 하면 1x,xx로 잘림수정
        except Exception as inst:
            print(salary)
        for i in range(4):
            #직급별  평균임금
            degree = driver.find_element_by_xpath('//*[@id="root"]/div/div[5]/div[1]/div[1]/div[2]/section[4]/div[2]/div[3]/div[2]/div['+str(i+1)+']/span').text
            salary=driver.find_element_by_xpath('//*[@id="root"]/div/div[5]/div[1]/div[1]/div[2]/section[4]/div[2]/div[3]/div[2]/div['+str(i+1)+']/div[2]').text
            salary = salary.replace(',', '')
            salary = salary[3:len(salary) - 2]
            print(degree,salary)
            ws.cell(row=r[0].row, column=i+5).value = int(salary)#1억이 넘는경우 3:8로 하면 1x,xx로 잘림수정
            if ws.cell(row=r[0].row, column=i + 5).value!=0:
                try:
                    ws.cell(row=r[0].row, column=i + 9).value=(ws.cell(row=r[0].row, column=i+5).value-ws.cell(row=r[0].row, column=i+4).value)/ws.cell(row=r[0].row, column=i+3).value
                except Exception as inst:
                    print(type(inst))
                    continue
            else:
                continue
        wb.save('1000companydataset.xlsx')
    else:
        continue
    print()
driver.quit()